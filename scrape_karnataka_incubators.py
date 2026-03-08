"""
Karnataka Startups Scraper - MeITY Startup Hub
================================================
Fetches ALL Karnataka startup listings from the MeITY Startup Hub API
(https://api.meity.gov.in) and exports detailed data including
real email, phone number, address, and social links to an Excel file.

REQUIREMENTS:
    pip install requests openpyxl pandas

USAGE:
    python scrape_karnataka_incubators.py

OUTPUT:
    karnataka_startups.xlsx  - Excel file with all startup details
"""

import time
import sys
import io
import requests
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fix Windows console encoding for Unicode
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# -----------------------------------------------
# API Configuration
# -----------------------------------------------

API_BASE = "https://api.meity.gov.in"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Origin": "https://msh.meity.gov.in",
    "Referer": "https://msh.meity.gov.in/",
}

# Concurrent workers for fetching detail pages (be polite but efficient)
MAX_WORKERS = 5
MAX_RETRIES = 3


# -----------------------------------------------
# Step 1 - Fetch all startups (list endpoint)
# -----------------------------------------------

def fetch_all_startups():
    """
    GET /startups?page_size=2000
    Paginates through all startups on the platform.
    """
    print("  [*] Fetching startup listing from API...")
    all_results = []
    page = 0
    page_size = 2000

    while True:
        r = requests.get(
            f"{API_BASE}/startups",
            headers=HEADERS,
            params={"page": page, "page_size": page_size},
            timeout=30,
        )
        r.raise_for_status()
        data = r.json()
        results = data.get("results", [])
        all_results.extend(results)
        print(f"    Page {page}: fetched {len(results)} startups (total so far: {len(all_results)})")
        if len(results) < page_size:
            break
        page += 1

    print(f"  [OK] Total startups across India: {len(all_results)}")
    return all_results


# -----------------------------------------------
# Step 2 - Filter for Karnataka
# -----------------------------------------------

def filter_karnataka(records):
    """Keep only records where state == Karnataka."""
    out = []
    for rec in records:
        state = (rec.get("state") or "").strip().lower()
        if state == "karnataka":
            out.append(rec)
    return out


# -----------------------------------------------
# Step 3 - Fetch full details for each startup
# -----------------------------------------------

def fetch_startup_detail(startup_id):
    """
    GET /startups/{id}
    Returns detailed startup info including full contact_info
    (email, phone, name, address, social links, etc.)
    """
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.get(
                f"{API_BASE}/startups/{startup_id}",
                headers=HEADERS,
                timeout=20,
            )
            if r.status_code == 200:
                data = r.json()
                results = data.get("results", [])
                if results:
                    return results[0]
            elif r.status_code == 429:
                wait = 2 ** attempt
                time.sleep(wait)
            else:
                return None
        except requests.RequestException:
            if attempt < MAX_RETRIES:
                time.sleep(1)
    return None


def fetch_detail_worker(args):
    """Worker function for thread pool."""
    idx, total, inc = args
    inc_id = inc.get("id")
    name = inc.get("name", "N/A")
    
    detail = fetch_startup_detail(inc_id) if inc_id else None
    
    if detail:
        record = build_record(detail)
        status = "OK"
    else:
        record = build_record(inc)
        status = "BASIC"
    
    return idx, name, record, status


def enrich_with_details(startups):
    """Fetch detail page for every startup using concurrent workers."""
    total = len(startups)
    print(f"\n  [*] Fetching detailed info for {total} Karnataka startups...")
    print(f"      Using {MAX_WORKERS} concurrent workers (this may take a few minutes)...")

    enriched = [None] * total
    tasks = [(i, total, s) for i, s in enumerate(startups)]
    
    completed = 0
    ok_count = 0
    basic_count = 0
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(fetch_detail_worker, t): t for t in tasks}
        
        for future in as_completed(futures):
            idx, name, record, status = future.result()
            enriched[idx] = record
            completed += 1
            
            if status == "OK":
                ok_count += 1
            else:
                basic_count += 1
            
            # Print progress every 25 startups
            if completed % 25 == 0 or completed == total:
                print(f"    Progress: {completed}/{total} ({ok_count} full, {basic_count} basic)")
    
    print(f"  [OK] Fetched details: {ok_count} full, {basic_count} basic only")
    return enriched


def build_record(data):
    """Build a flat dict from the API response for one startup."""
    contact = data.get("contact_info") or {}
    social = data.get("social_info") or {}

    # Contact info fields
    contact_name = contact.get("name", "")
    last_name = contact.get("lastName", "")
    full_name = f"{contact_name} {last_name}".strip() if contact_name or last_name else ""
    
    contact_email = contact.get("email", "")
    contact_email2 = contact.get("email2", "")
    contact_phone = contact.get("mobile", "") or contact.get("phone", "")
    contact_phone2 = contact.get("mobile2", "")
    
    # Address
    address_parts = []
    for key in ["address1", "address_line1", "address_line2", "address"]:
        val = contact.get(key) or data.get(key)
        if val and str(val).strip():
            address_parts.append(str(val).strip())
    address = ", ".join(address_parts) if address_parts else ""
    
    zip_code = contact.get("zipCode", "") or contact.get("zip_code", "")

    # Domain & sector - handle nested lists  
    domain = data.get("domain", [])
    if domain and isinstance(domain[0], list):
        domain = [item for sublist in domain for item in sublist]
    
    sector = data.get("sector", [])
    if sector and isinstance(sector[0], list):
        sector = [item for sublist in sector for item in sublist]

    return {
        "Name": data.get("name", ""),
        "City": data.get("city", ""),
        "State": data.get("state", "Karnataka"),
        "Zip Code": zip_code,
        "Address": address,
        "Contact Person": full_name,
        "Email": contact_email,
        "Email (Alt)": contact_email2,
        "Phone": contact_phone,
        "Phone (Alt)": contact_phone2,
        "Website": data.get("website_url", ""),
        "Startup Stage": data.get("startup_stage", ""),
        "Domain": ", ".join(domain) if isinstance(domain, list) else str(domain),
        "Sector": ", ".join(sector) if isinstance(sector, list) else str(sector),
        "Short Description": data.get("short_description", "") or "",
        "Incubator": data.get("incubator_name", ""),
        "Team Size": data.get("team_length", ""),
        "LinkedIn": social.get("linkedIn", ""),
        "Twitter": social.get("twitter", ""),
        "Facebook": social.get("facebook", ""),
        "Instagram": social.get("instagram", ""),
        "YouTube": social.get("youtube", ""),
    }


# -----------------------------------------------
# Step 4 - Excel export (beautifully formatted)
# -----------------------------------------------

COLUMNS = [
    "Name", "City", "State", "Zip Code", "Address",
    "Contact Person", "Email", "Email (Alt)", "Phone", "Phone (Alt)",
    "Website", "Startup Stage", "Domain", "Sector",
    "Short Description", "Incubator", "Team Size",
    "LinkedIn", "Twitter", "Facebook", "Instagram", "YouTube",
]


def export_to_excel(records, path="karnataka_startups.xlsx"):
    if not records:
        print("  [WARN] No records to export.")
        return

    df = pd.DataFrame(records, columns=COLUMNS)
    df.drop_duplicates(subset=["Name"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    # Handle case where file is open in Excel
    try:
        df.to_excel(path, index=False, sheet_name="Karnataka Startups")
    except PermissionError:
        import datetime
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        path = f"karnataka_startups_{ts}.xlsx"
        print(f"  [WARN] Original file is locked (maybe open in Excel).")
        print(f"         Saving to: {path}")
        df.to_excel(path, index=False, sheet_name="Karnataka Startups")

    # -- Formatting ----------------------------------------
    wb = load_workbook(path)
    ws = wb.active

    HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    EVEN_FILL   = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
    CELL_FONT   = Font(name="Arial", size=10)
    BORDER      = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    col_widths = {
        "Name": 35, "City": 18, "State": 14, "Zip Code": 10, "Address": 35,
        "Contact Person": 22, "Email": 32, "Email (Alt)": 32,
        "Phone": 16, "Phone (Alt)": 16,
        "Website": 30, "Startup Stage": 14, "Domain": 30, "Sector": 30,
        "Short Description": 50, "Incubator": 30, "Team Size": 10,
        "LinkedIn": 30, "Twitter": 25, "Facebook": 25,
        "Instagram": 25, "YouTube": 25,
    }

    for i, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i)
        cell.value = col
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 20)

    ws.row_dimensions[1].height = 30

    for row_idx in range(2, ws.max_row + 1):
        fill = EVEN_FILL if row_idx % 2 == 0 else None
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    print(f"\n  [OK] Saved -> {path}  ({len(df)} startups)")
    return path


# -----------------------------------------------
# MAIN
# -----------------------------------------------

def main():
    print("=" * 60)
    print("  Karnataka Startups Scraper - MeITY Startup Hub")
    print("=" * 60)

    # -- Step 1: Fetch all startups ----------------
    print("\n[1/4] Fetching startup list from API...")
    all_startups = fetch_all_startups()

    # -- Step 2: Filter for Karnataka ----------------
    print("\n[2/4] Filtering for Karnataka...")
    karnataka = filter_karnataka(all_startups)
    print(f"  [INFO] Karnataka startups found: {len(karnataka)}")

    if not karnataka:
        print("  [ERROR] No Karnataka startups found. Exiting.")
        return

    # -- Step 3: Fetch detailed info -----------------
    print("\n[3/4] Fetching detailed information for each startup...")
    print("       (email, phone, address, social links, etc.)")
    enriched = enrich_with_details(karnataka)

    # -- Step 4: Export to Excel ---------------------
    print("\n[4/4] Exporting to Excel...")
    export_to_excel(enriched)

    print("\n" + "=" * 60)
    print("  DONE! Open karnataka_startups.xlsx to see the data.")
    print("=" * 60)


if __name__ == "__main__":
    main()
