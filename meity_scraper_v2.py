"""
MeITY Startup Hub - India State Scraper v2
==========================================
Interactively fetches startup / incubator listings for ANY Indian state
from the MeITY Startup Hub API (https://api.meity.gov.in) and exports
detailed data to a formatted Excel file.

NEW IN v2:
    When "Startups" is selected, the script additionally asks whether you
    want Hardware startups, Software startups, or Both.
    * Hardware domains : Technology Hardware, Electronics, Embedded,
                         Semiconductor, Robotics, Drones
    * Software domains : Enterprise Software, IT Services, AI, Analytics,
                         AR VR (Augmented + Virtual Reality), Computer Vision,
                         Internet of Things, Security Solutions
    "Both" exports TWO separate xlsx files, one per category.

REQUIREMENTS:
    pip install requests openpyxl pandas

USAGE:
    python meity_scraper_v2.py
"""

import time
import sys
import io
import os
import datetime
import requests
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fix Windows console encoding for Unicode
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# -----------------------------------------------
# API Configuration
# -----------------------------------------------

API_BASE = "https://api.meity.gov.in"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Origin": "https://msh.meity.gov.in",
    "Referer": "https://msh.meity.gov.in/",
}

MAX_WORKERS = 5
MAX_RETRIES = 3

# -----------------------------------------------
# Hardware / Software domain classification
# -----------------------------------------------

HARDWARE_DOMAINS = {
    "technology hardware",
    "electronics",
    "embedded",
    "semiconductor",
    "robotics",
    "drones",
}

SOFTWARE_DOMAINS = {
    "enterprise software",
    "it services",
    "ai",
    "analytics",
    "ar vr (augmented + virtual reality)",
    "computer vision",
    "internet of things",
    "security solutions",
}

# -----------------------------------------------
# Known Indian States & UTs
# -----------------------------------------------

KNOWN_STATES = [
    "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh",
    "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jharkhand",
    "Karnataka", "Kerala", "Madhya Pradesh", "Maharashtra", "Manipur",
    "Meghalaya", "Mizoram", "Nagaland", "Odisha", "Punjab",
    "Rajasthan", "Sikkim", "Tamil Nadu", "Telangana", "Tripura",
    "Uttar Pradesh", "Uttarakhand", "West Bengal",
    "Andaman and Nicobar Islands", "Chandigarh",
    "Dadra and Nagar Haveli and Daman and Diu",
    "Delhi", "Jammu and Kashmir", "Ladakh", "Lakshadweep", "Puducherry",
]

ENTITY_ENDPOINTS = {
    "startups":   "/startups",
    "incubators": "/incubators",
}


# -----------------------------------------------
# Step 0 - Interactive User Input
# -----------------------------------------------

def ask_entity_type():
    print("\nWhat do you want to fetch?")
    print("  [1] Startups")
    print("  [2] Incubators")
    while True:
        choice = input("\nEnter 1 or 2 (default: 1): ").strip() or "1"
        if choice == "1":
            return "startups"
        elif choice == "2":
            return "incubators"
        else:
            print("  [!] Please enter 1 or 2.")


def ask_startup_category():
    """Ask Hardware / Software / Both when Startups is selected."""
    print("\nWhat type of startups do you want?")
    print("  [1] Hardware  (Technology Hardware, Electronics, Embedded,")
    print("                  Semiconductor, Robotics, Drones)")
    print("  [2] Software  (Enterprise Software, IT Services, AI, Analytics,")
    print("                  AR VR, Computer Vision, IoT, Security Solutions)")
    print("  [3] Both      (exports two separate xlsx files)")
    while True:
        choice = input("\nEnter 1, 2, or 3 (default: 1): ").strip() or "1"
        if choice == "1":
            return "hardware"
        elif choice == "2":
            return "software"
        elif choice == "3":
            return "both"
        else:
            print("  [!] Please enter 1, 2, or 3.")


def ask_state():
    print("\nAvailable states / UTs:")
    for i, s in enumerate(KNOWN_STATES, 1):
        end = "\n" if i % 4 == 0 else "\t"
        print(f"  {s}", end=end)
    print()
    while True:
        raw = input("\nEnter state name (e.g. Karnataka): ").strip()
        if not raw:
            print("  [!] State name cannot be empty.")
            continue
        match = next((s for s in KNOWN_STATES if s.lower() == raw.lower()), None)
        if match:
            return match
        else:
            print(f"  [?] '{raw}' is not in the known list.")
            confirm = input("    Use it anyway? (y/n): ").strip().lower()
            if confirm == "y":
                return raw.title()


# -----------------------------------------------
# Step 1 - Fetch all entities from API
# -----------------------------------------------

def fetch_all_entities(entity_type):
    endpoint = ENTITY_ENDPOINTS[entity_type]
    print(f"  [*] Fetching {entity_type} listing from API...")
    all_results = []
    page = 0
    page_size = 2000
    while True:
        r = requests.get(
            f"{API_BASE}{endpoint}",
            headers=HEADERS,
            params={"page": page, "page_size": page_size},
            timeout=30,
        )
        r.raise_for_status()
        data = r.json()
        results = data.get("results", [])
        all_results.extend(results)
        print(f"    Page {page}: fetched {len(results)} {entity_type} (total: {len(all_results)})")
        if len(results) < page_size:
            break
        page += 1
    print(f"  [OK] Total {entity_type} across India: {len(all_results)}")
    return all_results


# -----------------------------------------------
# Step 2 - Filter by state
# -----------------------------------------------

def filter_by_state(records, state):
    target = state.strip().lower()
    return [r for r in records if (r.get("state") or "").strip().lower() == target]


# -----------------------------------------------
# Step 2b - Filter startups by Hardware / Software
# -----------------------------------------------

def _get_domains_for_record(record):
    domain = record.get("domain", [])
    if not domain:
        return set()
    if isinstance(domain, list):
        flat = []
        for item in domain:
            if isinstance(item, list):
                flat.extend(item)
            else:
                flat.append(item)
        return {str(d).strip().lower() for d in flat if d}
    return {str(domain).strip().lower()}


def filter_by_category(records, category):
    target_set = HARDWARE_DOMAINS if category == "hardware" else SOFTWARE_DOMAINS
    return [r for r in records if _get_domains_for_record(r) & target_set]


# -----------------------------------------------
# Step 3 - Fetch full details
# -----------------------------------------------

def fetch_entity_detail(entity_id, entity_type):
    endpoint = ENTITY_ENDPOINTS[entity_type]
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.get(f"{API_BASE}{endpoint}/{entity_id}", headers=HEADERS, timeout=20)
            if r.status_code == 200:
                data = r.json()
                results = data.get("results", [])
                if results:
                    return results[0]
            elif r.status_code == 429:
                time.sleep(2 ** attempt)
            else:
                return None
        except requests.RequestException:
            if attempt < MAX_RETRIES:
                time.sleep(1)
    return None


def fetch_detail_worker(args):
    idx, total, record, entity_type = args
    eid = record.get("id")
    detail = fetch_entity_detail(eid, entity_type) if eid else None
    if detail:
        built = build_record(detail)
        status = "OK"
    else:
        built = build_record(record)
        status = "BASIC"
    return idx, record.get("name", "N/A"), built, status


def enrich_with_details(entities, state, entity_type):
    total = len(entities)
    print(f"\n  [*] Fetching detailed info for {total} {state} {entity_type}...")
    print(f"      Using {MAX_WORKERS} concurrent workers...")
    enriched = [None] * total
    tasks = [(i, total, e, entity_type) for i, e in enumerate(entities)]
    completed = ok_count = basic_count = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(fetch_detail_worker, t): t for t in tasks}
        for future in as_completed(futures):
            idx, name, record, status = future.result()
            enriched[idx] = record
            completed += 1
            if status == "OK":
                ok_count += 1
            else:
                basic_count += 1
            if completed % 25 == 0 or completed == total:
                print(f"    Progress: {completed}/{total} ({ok_count} full, {basic_count} basic)")
    print(f"  [OK] Fetched details: {ok_count} full, {basic_count} basic only")
    return enriched


def _first(*values):
    for v in values:
        if v and str(v).strip():
            return str(v).strip()
    return ""


def build_record(data):
    contact = data.get("contact_info") or {}
    social  = data.get("social_info")  or {}

    contact_name = _first(contact.get("name"), data.get("contact_name"),
                          data.get("poc_name"), data.get("spoc_name"))
    last_name    = _first(contact.get("lastName"), data.get("last_name"))
    full_name    = f"{contact_name} {last_name}".strip()

    contact_email  = _first(contact.get("email"), data.get("email"),
                            data.get("contact_email"), data.get("poc_email"), data.get("spoc_email"))
    contact_email2 = _first(contact.get("email2"), data.get("email2"), data.get("alternate_email"))

    contact_phone  = _first(contact.get("mobile"), contact.get("phone"),
                            data.get("mobile"), data.get("phone"),
                            data.get("contact_mobile"), data.get("contact_phone"),
                            data.get("poc_mobile"), data.get("spoc_mobile"))
    contact_phone2 = _first(contact.get("mobile2"), data.get("mobile2"), data.get("alternate_phone"))

    address_parts = []
    for key in ["address1", "address_line1", "address_line2", "address"]:
        val = contact.get(key) or data.get(key)
        if val and str(val).strip():
            address_parts.append(str(val).strip())
    address = ", ".join(address_parts) if address_parts else ""

    zip_code = _first(contact.get("zipCode"), contact.get("zip_code"),
                      data.get("zipCode"), data.get("zip_code"), data.get("pincode"))

    domain = data.get("domain", [])
    if domain and isinstance(domain, list) and domain and isinstance(domain[0], list):
        domain = [item for sublist in domain for item in sublist]

    sector = data.get("sector", [])
    if sector and isinstance(sector, list) and sector and isinstance(sector[0], list):
        sector = [item for sublist in sector for item in sublist]

    website = _first(data.get("website_url"), data.get("website"), data.get("web_url"))

    return {
        "Name":              data.get("name", ""),
        "City":              data.get("city", ""),
        "State":             data.get("state", ""),
        "Zip Code":          zip_code,
        "Address":           address,
        "Contact Person":    full_name,
        "Email":             contact_email,
        "Email (Alt)":       contact_email2,
        "Phone":             contact_phone,
        "Phone (Alt)":       contact_phone2,
        "Website":           website,
        "Startup Stage":     data.get("startup_stage", ""),
        "Domain":            ", ".join(domain) if isinstance(domain, list) else str(domain),
        "Sector":            ", ".join(sector) if isinstance(sector, list) else str(sector),
        "Short Description": data.get("short_description", "") or "",
        "Incubator":         data.get("incubator_name", ""),
        "Team Size":         data.get("team_length", ""),
        "LinkedIn":          _first(social.get("linkedIn"),  data.get("linkedin")),
        "Twitter":           _first(social.get("twitter"),   data.get("twitter")),
        "Facebook":          _first(social.get("facebook"),  data.get("facebook")),
        "Instagram":         _first(social.get("instagram"), data.get("instagram")),
        "YouTube":           _first(social.get("youtube"),   data.get("youtube")),
    }


# -----------------------------------------------
# Step 4 - Excel export
# -----------------------------------------------

COLUMNS = [
    "Name", "City", "State", "Zip Code", "Address",
    "Contact Person", "Email", "Email (Alt)", "Phone", "Phone (Alt)",
    "Website", "Startup Stage", "Domain", "Sector",
    "Short Description", "Incubator", "Team Size",
    "LinkedIn", "Twitter", "Facebook", "Instagram", "YouTube",
]

COL_WIDTHS = {
    "Name": 35, "City": 18, "State": 14, "Zip Code": 10, "Address": 35,
    "Contact Person": 22, "Email": 32, "Email (Alt)": 32,
    "Phone": 16, "Phone (Alt)": 16,
    "Website": 30, "Startup Stage": 14, "Domain": 30, "Sector": 30,
    "Short Description": 50, "Incubator": 30, "Team Size": 10,
    "LinkedIn": 30, "Twitter": 25, "Facebook": 25,
    "Instagram": 25, "YouTube": 25,
}

HEADER_COLORS = {
    "hardware":   "7B3F00",
    "software":   "1F4E79",
    "incubators": "14532D",
}

EVEN_FILLS = {
    "hardware":   "F5E6D3",
    "software":   "D6E4F0",
    "incubators": "D1FAE5",
}


def export_to_excel(records, state, entity_type, category=None):
    if not records:
        label = category if category else entity_type
        print(f"  [WARN] No records to export for {label}.")
        return None

    folder = entity_type
    os.makedirs(folder, exist_ok=True)

    safe_state = state.lower().replace(" ", "_")

    if category:
        filename   = f"{safe_state}_{category}_startups.xlsx"
        sheet_name = f"{state} {category.title()} Startups"[:31]
        color_key  = category
    else:
        filename   = f"{safe_state}_{entity_type}.xlsx"
        sheet_name = f"{state} {entity_type.title()}"[:31]
        color_key  = entity_type

    path = os.path.join(folder, filename)

    df = pd.DataFrame(records, columns=COLUMNS)
    df.drop_duplicates(subset=["Name"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    try:
        df.to_excel(path, index=False, sheet_name=sheet_name)
    except PermissionError:
        ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base = filename.replace(".xlsx", "")
        path = os.path.join(folder, f"{base}_{ts}.xlsx")
        print(f"  [WARN] File locked, saving to: {path}")
        df.to_excel(path, index=False, sheet_name=sheet_name)

    wb = load_workbook(path)
    ws = wb.active

    hdr_color  = HEADER_COLORS.get(color_key, "1F4E79")
    even_color = EVEN_FILLS.get(color_key, "D6E4F0")

    HEADER_FILL = PatternFill("solid", start_color=hdr_color,  end_color=hdr_color)
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    EVEN_FILL   = PatternFill("solid", start_color=even_color, end_color=even_color)
    CELL_FONT   = Font(name="Arial", size=10)
    BORDER      = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    for i, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i)
        cell.value = col
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(col, 20)

    ws.row_dimensions[1].height = 30

    for row_idx in range(2, ws.max_row + 1):
        fill = EVEN_FILL if row_idx % 2 == 0 else None
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font   = CELL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    print(f"\n  [OK] Saved -> {path}  ({len(df)} records)")
    return path


# -----------------------------------------------
# MAIN
# -----------------------------------------------

def main():
    print("=" * 60)
    print("  MeITY Startup Hub - India State Scraper  v2")
    print("=" * 60)

    entity_type = ask_entity_type()

    startup_category = None
    if entity_type == "startups":
        startup_category = ask_startup_category()

    state = ask_state()

    print(f"\n  >> Fetching {entity_type} for: {state}\n")

    print("[1/4] Fetching full listing from API...")
    all_entities = fetch_all_entities(entity_type)

    print(f"\n[2/4] Filtering for {state}...")
    state_filtered = filter_by_state(all_entities, state)
    print(f"  [INFO] {state} {entity_type} found: {len(state_filtered)}")

    if not state_filtered:
        print(f"  [ERROR] No {entity_type} found for '{state}'. Check the state name and try again.")
        return

    if entity_type == "startups" and startup_category:
        categories_to_export = (
            ["hardware", "software"] if startup_category == "both" else [startup_category]
        )

        category_buckets = {}
        for cat in categories_to_export:
            bucket = filter_by_category(state_filtered, cat)
            category_buckets[cat] = bucket
            print(f"  [INFO] {state} {cat.title()} startups matched: {len(bucket)}")

        print(f"\n[3/4] Fetching detailed information...")
        print("       (email, phone, address, social links, etc.)")

        saved_paths = []
        for cat, bucket in category_buckets.items():
            if not bucket:
                print(f"  [SKIP] No {cat} startups found for {state}.")
                continue
            print(f"\n  --- Enriching {len(bucket)} {cat.title()} startups ---")
            enriched = enrich_with_details(bucket, state, entity_type)
            print(f"\n[4/4] Exporting {cat.title()} startups to Excel...")
            out_path = export_to_excel(enriched, state, entity_type, category=cat)
            if out_path:
                saved_paths.append(out_path)

        print("\n" + "=" * 60)
        if saved_paths:
            print("  DONE! Files saved:")
            for p in saved_paths:
                print(f"    -> {p}")
        print("=" * 60)

    else:
        print(f"\n[3/4] Fetching detailed information for each {entity_type.rstrip('s')}...")
        print("       (email, phone, address, social links, etc.)")
        enriched = enrich_with_details(state_filtered, state, entity_type)

        print("\n[4/4] Exporting to Excel...")
        out_path = export_to_excel(enriched, state, entity_type)

        print("\n" + "=" * 60)
        if out_path:
            print(f"  DONE! Open {out_path} to see the data.")
        print("=" * 60)


if __name__ == "__main__":
    main()
